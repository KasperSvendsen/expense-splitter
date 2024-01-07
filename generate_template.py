import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import openpyxl

def get_document_name():
    return input("Enter the name of the document: ")

def get_group_size():
    while True:
        try:
            size = int(input("Enter the number of people in the group: "))
            if size > 0:
                return size
            else:
                print("Please enter a valid positive number.")
        except ValueError:
            print("Please enter a valid number.")

def get_person_names(size):
    names = []
    for i in range(size):
        name = input(f"Enter the name of person {i + 1}: ")
        names.append(name)
    return names

def get_default_currency():
    return input("Enter the default currency code (e.g. DKK or EUR): ")

def create_expense_template(doc_name, people_names, default_currency):
    data = {'Paying person': [], 'Description': [], 'Amount': [], 'Currency': [], 'Shared with': []}
    
    # Initialize columns for each person's share
    for person in people_names:
        data[f"{person}'s share"] = []

    # Number of people in the group
    num_people = len(people_names)

    # Add 10 blank lines with default values/formulas
    for i in range(10):
        data['Paying person'].append('')
        data['Description'].append('')
        data['Amount'].append('')
        data['Currency'].append(default_currency)
        data['Shared with'].append(', '.join(people_names))

         # Adding share formula for each person
        for person in people_names:
            # Formula to check if person is included in 'Shared with' and calculate share
            data[f"{person}'s share"].append(f'=IF(ISNUMBER(SEARCH("{person}"; E{i+2})); C{i+2} / (LEN(E{i+2})-LEN(SUBSTITUTE(E{i+2}; ","; ""))+1); 0)')
    
    # Create DataFrame from the base data
    df = pd.DataFrame(data)

     # Create a new Workbook and add the DataFrame to it
    book = Workbook()
    book.remove(book.active)
    df.to_excel(f"{doc_name}.xlsx", index=False, engine='openpyxl')

    # Open the workbook again using openpyxl for further modifications
    book = load_workbook(f"{doc_name}.xlsx")
    sheet = book.active

    # Add dropdown list for "Paying person" column
    dv = DataValidation(type="list", formula1=f'"{", ".join(people_names)}"', allow_blank=True)
    sheet.add_data_validation(dv)
    dv.add(f'A2:A{len(people_names)+1}')  # Assuming A is the column for "Paying person"

    # Find maximum length of the share header
    max_share_header_length = max(len(f"{person}'s share") for person in people_names)

    # Set the width of the columns, including the new columns
    for i, column_cells in enumerate(sheet.columns):
        column_letter = openpyxl.utils.get_column_letter(i+1)
        if i >= 5:  # This assumes that the first share column is 'F'
            sheet.column_dimensions[column_letter].width = max_share_header_length + 2
        else:
            max_length = max(len(str(cell.value)) for cell in column_cells if cell.value)
            adjusted_width = max_length + 2
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Save the Excel file
    book.save(f"{doc_name}.xlsx")
    print(f"Expense template '{doc_name}.xlsx' created successfully!")


def main():
    document_name = get_document_name()
    group_size = get_group_size()
    people_names = get_person_names(group_size)
    default_currency = get_default_currency()
    create_expense_template(document_name, people_names, default_currency)

if __name__ == "__main__":
    main()
